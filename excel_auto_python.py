import openpyxl
import openpyxl.worksheet.worksheet
from openpyxl import utils


class excel_automation_python:

    def __init__(self, file_path_with_name):
        self.file_path_with_name = file_path_with_name

    def load_file(self):
        xl = openpyxl.load_workbook(self.file_path_with_name)
        return xl

    def copy_past_column_based_on_text_value(self, sheet_from, sheet_to, column_from, column_to, search_criteria):

        xl = openpyxl.load_workbook(self.file_path_with_name)
        sheet1 = xl[sheet_from]
        sheet2 = xl[sheet_to]
        lastrow = sheet1.max_row + 1
        rangeselect = []
        rangeselect.append(sheet1.cell(row=1, column=column_from).value)
        for i in range(1, lastrow, 1):
            if sheet1.cell(row=i, column=column_from).value == search_criteria:
                rangeselect.extend([sheet1.cell(row=i, column=column_from).value])
        for i in range(1, len(rangeselect) + 1, 1):
            sheet2.cell(row=i, column=column_to).value = rangeselect[i - 1]
        xl.save(self.file_path_with_name)

    def copy_paste_based_on_field_heading(self, sheet_from, sheet_to, heading_names):

        xl = openpyxl.load_workbook(self.file_path_with_name)
        sheet1 = xl[sheet_from]
        sheet2 = xl[sheet_to]
        last_column = sheet1.max_column
        last_column2 = sheet2.max_column
        heading_names = list(heading_names)

        for i in range(1, last_column + 1, 1):
            for heading_name in heading_names:
                if sheet1.cell(row=1, column=i).value == heading_name:
                    column_index = utils.get_column_letter(i)
                    if sheet2.cell(row=1, column=1).value is None:
                        for cell in sheet1[column_index]:
                            sheet2.cell(row=cell.row, column=last_column2, value=cell.value)
                    else:
                        for cell in sheet1[column_index]:
                            sheet2.cell(row=cell.row, column=last_column2 + 1, value=cell.value)

        xl.save(self.file_path_with_name)

    def create_formula(self, sheet_name, cell_ref, formula):
        xl = openpyxl.load_workbook(self.file_path_with_name)
        sheet1 = xl[sheet_name]
        sheet1[cell_ref] = formula

        xl.save(self.file_path_with_name)
